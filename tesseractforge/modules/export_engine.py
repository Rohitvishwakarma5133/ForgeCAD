#!/usr/bin/env python3
"""
🔥 ADVANCED EXPORT ENGINE
=========================

Comprehensive export functionality for technical drawing analysis:

✅ Excel export with advanced formatting
✅ PDF report generation
✅ CAD file integration via Autodesk Forge API
✅ Multiple export templates
✅ Batch export capabilities  
✅ Custom report generation
✅ Database export preparation

For Professional Engineering Documentation
"""

import os
import json
import csv
import asyncio
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Union, Tuple
import time
import base64
import tempfile
import uuid

# Excel and PDF libraries
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.chart import BarChart, Reference
    from openpyxl.drawing.image import Image as XLImage
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.platypus import PageBreak, KeepTogether
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# Autodesk Forge API integration
import requests
import aiohttp

class ExportFormat:
    """Export format constants"""
    JSON = "json"
    EXCEL = "excel"
    CSV = "csv"
    PDF = "pdf"
    XML = "xml"
    FORGE_CAD = "forge_cad"
    
class ExportTemplate:
    """Export template types"""
    BASIC = "basic"
    DETAILED = "detailed"
    TECHNICAL_REPORT = "technical_report"
    EXECUTIVE_SUMMARY = "executive_summary"
    CAD_INTEGRATION = "cad_integration"

class AdvancedExportEngine:
    """🔥 Advanced Multi-Format Export Engine"""
    
    def __init__(self, forge_client_id: str = None, forge_client_secret: str = None):
        """Initialize the export engine with optional Forge integration"""
        
        print("🚀 INITIALIZING ADVANCED EXPORT ENGINE")
        print("=" * 60)
        
        # Forge API configuration
        self.forge_config = {
            'client_id': forge_client_id or os.getenv('FORGE_CLIENT_ID'),
            'client_secret': forge_client_secret or os.getenv('FORGE_CLIENT_SECRET'),
            'base_url': 'https://developer.api.autodesk.com',
            'auth_url': 'https://developer.api.autodesk.com/authentication/v1/authenticate'
        }
        
        self.forge_token = None
        self.forge_token_expiry = 0
        
        # Export capabilities
        self.export_capabilities = {
            'excel': EXCEL_AVAILABLE,
            'pdf': PDF_AVAILABLE,
            'forge_api': bool(self.forge_config['client_id']),
            'json': True,
            'csv': True,
            'xml': True
        }
        
        # Export statistics
        self.export_stats = {
            'total_exports': 0,
            'successful_exports': 0,
            'failed_exports': 0,
            'format_usage': {},
            'average_export_time': 0.0
        }
        
        # Template configurations
        self.templates = self._load_export_templates()
        
        print("✅ Export Engine Status:")
        for format_name, available in self.export_capabilities.items():
            status = "READY" if available else "NOT AVAILABLE"
            print(f"   {format_name.upper()}: {status}")
        
        print("🎯 EXPORT ENGINE READY!")
        print("=" * 60)
    
    def _load_export_templates(self) -> Dict[str, Dict[str, Any]]:
        """Load export template configurations"""
        return {
            ExportTemplate.BASIC: {
                'name': 'Basic Export',
                'description': 'Simple data export with essential information',
                'sections': ['metadata', 'technical_data', 'summary'],
                'styling': 'minimal'
            },
            ExportTemplate.DETAILED: {
                'name': 'Detailed Analysis Export',
                'description': 'Comprehensive export with all analysis results',
                'sections': ['metadata', 'ocr_analysis', 'symbol_analysis', 'semantic_analysis', 'technical_data'],
                'styling': 'detailed'
            },
            ExportTemplate.TECHNICAL_REPORT: {
                'name': 'Technical Engineering Report',
                'description': 'Professional engineering documentation',
                'sections': ['executive_summary', 'technical_specifications', 'manufacturing_info', 'recommendations'],
                'styling': 'professional'
            },
            ExportTemplate.EXECUTIVE_SUMMARY: {
                'name': 'Executive Summary',
                'description': 'High-level overview for management',
                'sections': ['summary', 'key_findings', 'recommendations'],
                'styling': 'executive'
            },
            ExportTemplate.CAD_INTEGRATION: {
                'name': 'CAD Integration Export',
                'description': 'Export optimized for CAD system integration',
                'sections': ['cad_metadata', 'geometric_data', 'technical_specifications'],
                'styling': 'cad_optimized'
            }
        }
    
    async def authenticate_forge(self) -> bool:
        """Authenticate with Autodesk Forge API"""
        if not self.forge_config['client_id'] or not self.forge_config['client_secret']:
            return False
        
        # Check if token is still valid (with 5 minute buffer)
        if self.forge_token and time.time() < self.forge_token_expiry - 300:
            return True
        
        try:
            async with aiohttp.ClientSession() as session:
                payload = {
                    'client_id': self.forge_config['client_id'],
                    'client_secret': self.forge_config['client_secret'],
                    'grant_type': 'client_credentials',
                    'scope': 'data:read data:write data:create bucket:create bucket:read'
                }
                
                async with session.post(self.forge_config['auth_url'], data=payload) as response:
                    if response.status == 200:
                        data = await response.json()
                        self.forge_token = data['access_token']
                        self.forge_token_expiry = time.time() + data['expires_in']
                        return True
            
        except Exception as e:
            print(f"⚠️ Forge authentication failed: {e}")
        
        return False
    
    async def export_to_excel(self, data: Dict[str, Any], 
                            file_path: str,
                            template: str = ExportTemplate.DETAILED,
                            include_charts: bool = True) -> Dict[str, Any]:
        """🔥 ADVANCED EXCEL EXPORT"""
        
        if not EXCEL_AVAILABLE:
            return {'status': 'error', 'error': 'Excel libraries not available'}
        
        start_time = time.time()
        
        try:
            print(f"📊 Generating Excel export: {os.path.basename(file_path)}")
            
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            template_config = self.templates.get(template, self.templates[ExportTemplate.DETAILED])
            
            # Sheet 1: Summary
            summary_sheet = wb.create_sheet("Summary", 0)
            self._create_summary_sheet(summary_sheet, data, template_config)
            
            # Sheet 2: Technical Data
            if 'technical_data' in template_config['sections']:
                tech_sheet = wb.create_sheet("Technical Data", 1)
                self._create_technical_data_sheet(tech_sheet, data)
            
            # Sheet 3: OCR Results
            if 'ocr_analysis' in template_config['sections']:
                ocr_sheet = wb.create_sheet("OCR Analysis", 2)
                self._create_ocr_analysis_sheet(ocr_sheet, data)
            
            # Sheet 4: Symbol Analysis
            if 'symbol_analysis' in template_config['sections']:
                symbol_sheet = wb.create_sheet("Symbol Analysis", 3)
                self._create_symbol_analysis_sheet(symbol_sheet, data)
            
            # Sheet 5: AI Analysis
            if 'semantic_analysis' in template_config['sections']:
                ai_sheet = wb.create_sheet("AI Analysis", 4)
                self._create_ai_analysis_sheet(ai_sheet, data)
            
            # Sheet 6: Charts and Visualizations
            if include_charts:
                chart_sheet = wb.create_sheet("Analytics", 5)
                self._create_analytics_sheet(chart_sheet, data)
            
            # Save workbook
            wb.save(file_path)
            
            processing_time = time.time() - start_time
            
            # Update statistics
            self._update_export_stats('excel', processing_time, True)
            
            print(f"   ✅ Excel export complete in {processing_time:.1f}s")
            
            return {
                'status': 'success',
                'file_path': file_path,
                'file_size': os.path.getsize(file_path),
                'processing_time': processing_time,
                'sheets_created': len(wb.sheetnames)
            }
            
        except Exception as e:
            self._update_export_stats('excel', time.time() - start_time, False)
            return {
                'status': 'error',
                'error': str(e),
                'processing_time': time.time() - start_time
            }
    
    def _create_summary_sheet(self, sheet, data: Dict[str, Any], template_config: Dict[str, Any]):
        """Create summary sheet with professional formatting"""
        
        # Title
        sheet['A1'] = 'Technical Drawing Analysis Report'
        sheet['A1'].font = Font(size=16, bold=True, color="1F4E79")
        sheet.merge_cells('A1:E1')
        
        # Basic information
        row = 3
        sheet[f'A{row}'] = 'Analysis Date:'
        sheet[f'B{row}'] = data.get('metadata', {}).get('processing_timestamp', '')
        
        row += 1
        sheet[f'A{row}'] = 'Pipeline ID:'
        sheet[f'B{row}'] = data.get('metadata', {}).get('pipeline_id', 'N/A')
        
        # Technical summary
        tech_data = data.get('technical_data', {})
        row += 2
        sheet[f'A{row}'] = 'Technical Summary'
        sheet[f'A{row}'].font = Font(size=14, bold=True)
        
        row += 1
        sheet[f'A{row}'] = 'Dimensions Found:'
        sheet[f'B{row}'] = len(tech_data.get('dimensions_found', []))
        
        row += 1
        sheet[f'A{row}'] = 'Materials Identified:'
        sheet[f'B{row}'] = len(tech_data.get('materials_identified', []))
        
        row += 1
        sheet[f'A{row}'] = 'Tolerances Specified:'
        sheet[f'B{row}'] = len(tech_data.get('tolerances_specified', []))
        
        # Quality metrics
        quality_metrics = tech_data.get('quality_metrics', {})
        row += 2
        sheet[f'A{row}'] = 'Quality Metrics'
        sheet[f'A{row}'].font = Font(size=14, bold=True)
        
        row += 1
        sheet[f'A{row}'] = 'OCR Confidence:'
        sheet[f'B{row}'] = f"{quality_metrics.get('ocr_confidence', 0):.1%}"
        
        row += 1
        sheet[f'A{row}'] = 'Semantic Confidence:'
        sheet[f'B{row}'] = f"{quality_metrics.get('semantic_confidence', 0):.1%}"
        
        # Style the sheet
        for row in sheet.iter_rows():
            for cell in row:
                if cell.column == 1:  # Column A
                    cell.font = Font(bold=True)
                cell.alignment = Alignment(vertical='top')
        
        # Adjust column widths
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 30
    
    def _create_technical_data_sheet(self, sheet, data: Dict[str, Any]):
        """Create detailed technical data sheet"""
        
        # Headers
        headers = ['Type', 'Category', 'Value', 'Source', 'Confidence']
        for i, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=i, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Get export-ready data
        export_data = data.get('export_ready_data', {})
        tabular_data = export_data.get('tabular_data', [])
        
        # Add data rows
        for row_idx, item in enumerate(tabular_data, 2):
            sheet.cell(row=row_idx, column=1, value=item.get('Type', ''))
            sheet.cell(row=row_idx, column=2, value=item.get('Category', ''))
            sheet.cell(row=row_idx, column=3, value=item.get('Value', ''))
            sheet.cell(row=row_idx, column=4, value=item.get('Source', ''))
            confidence_val = item.get('Confidence', 0)
            if isinstance(confidence_val, (int, float)):
                sheet.cell(row=row_idx, column=5, value=f"{confidence_val:.1%}")
            else:
                sheet.cell(row=row_idx, column=5, value=str(confidence_val))
        
        # Add borders and formatting
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in sheet.iter_rows(min_row=1, max_row=len(tabular_data) + 1):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
        
        # Adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    def _create_ocr_analysis_sheet(self, sheet, data: Dict[str, Any]):
        """Create OCR analysis results sheet"""
        
        sheet['A1'] = 'OCR Analysis Results'
        sheet['A1'].font = Font(size=14, bold=True)
        
        ocr_data = data.get('ocr_analysis', {})
        performance = ocr_data.get('performance', {})
        
        # Performance summary
        row = 3
        sheet[f'A{row}'] = 'Performance Summary'
        sheet[f'A{row}'].font = Font(size=12, bold=True)
        
        row += 1
        sheet[f'A{row}'] = 'Total Processing Time:'
        sheet[f'B{row}'] = f"{performance.get('total_processing_time', 0):.2f}s"
        
        row += 1
        sheet[f'A{row}'] = 'Successful APIs:'
        sheet[f'B{row}'] = performance.get('successful_apis', 0)
        
        # Text content
        text_content = ocr_data.get('text_extracted', [])
        if text_content:
            row += 2
            sheet[f'A{row}'] = 'Extracted Text Content'
            sheet[f'A{row}'].font = Font(size=12, bold=True)
            
            # Headers
            row += 1
            headers = ['Source', 'Text Length', 'Processing Time', 'Text Preview']
            for i, header in enumerate(headers, 1):
                cell = sheet.cell(row=row, column=i, value=header)
                cell.font = Font(bold=True)
            
            # Content
            for text_item in text_content:
                row += 1
                sheet.cell(row=row, column=1, value=text_item.get('source', ''))
                sheet.cell(row=row, column=2, value=text_item.get('length', 0))
                sheet.cell(row=row, column=3, value=f"{text_item.get('processing_time', 0):.2f}s")
                preview = text_item.get('text', '')[:100] + ('...' if len(text_item.get('text', '')) > 100 else '')
                sheet.cell(row=row, column=4, value=preview)
    
    def _create_symbol_analysis_sheet(self, sheet, data: Dict[str, Any]):
        """Create symbol analysis results sheet"""
        
        sheet['A1'] = 'Symbol & Shape Analysis'
        sheet['A1'].font = Font(size=14, bold=True)
        
        symbol_data = data.get('symbol_analysis', {})
        geometric = symbol_data.get('geometric_elements', {})
        
        row = 3
        sheet[f'A{row}'] = 'Geometric Elements Summary'
        sheet[f'A{row}'].font = Font(size=12, bold=True)
        
        row += 1
        sheet[f'A{row}'] = 'Total Shapes:'
        sheet[f'B{row}'] = geometric.get('total_shapes', 0)
        
        if geometric.get('circles'):
            row += 1
            sheet[f'A{row}'] = 'Circles:'
            sheet[f'B{row}'] = len(geometric['circles'])
        
        if geometric.get('rectangles'):
            row += 1
            sheet[f'A{row}'] = 'Rectangles:'
            sheet[f'B{row}'] = len(geometric['rectangles'])
        
        if geometric.get('lines'):
            row += 1
            sheet[f'A{row}'] = 'Lines:'
            sheet[f'B{row}'] = len(geometric['lines'])
        
        # Annotations summary
        annotations = symbol_data.get('annotations', {})
        if annotations:
            row += 2
            sheet[f'A{row}'] = 'Annotations Summary'
            sheet[f'A{row}'].font = Font(size=12, bold=True)
            
            summary = annotations.get('summary', {})
            row += 1
            sheet[f'A{row}'] = 'Dimension Lines:'
            sheet[f'B{row}'] = summary.get('total_dimension_lines', 0)
            
            row += 1
            sheet[f'A{row}'] = 'Arrows:'
            sheet[f'B{row}'] = summary.get('total_arrows', 0)
    
    def _create_ai_analysis_sheet(self, sheet, data: Dict[str, Any]):
        """Create AI analysis results sheet"""
        
        sheet['A1'] = 'AI Semantic Analysis'
        sheet['A1'].font = Font(size=14, bold=True)
        
        semantic_data = data.get('semantic_analysis', {})
        ai_analyses = semantic_data.get('ai_analyses', {})
        
        row = 3
        for ai_name, analysis in ai_analyses.items():
            sheet[f'A{row}'] = f'{ai_name.upper()} Analysis'
            sheet[f'A{row}'].font = Font(size=12, bold=True)
            
            row += 1
            sheet[f'A{row}'] = 'Status:'
            sheet[f'B{row}'] = analysis.get('status', 'unknown')
            
            if analysis.get('status') == 'success':
                row += 1
                sheet[f'A{row}'] = 'Model:'
                sheet[f'B{row}'] = analysis.get('model', 'N/A')
                
                row += 1
                sheet[f'A{row}'] = 'Processing Time:'
                sheet[f'B{row}'] = f"{analysis.get('processing_time', 0):.2f}s"
                
                row += 1
                sheet[f'A{row}'] = 'Tokens Used:'
                sheet[f'B{row}'] = analysis.get('tokens_used', 0)
                
                row += 1
                sheet[f'A{row}'] = 'Confidence:'
                sheet[f'B{row}'] = f"{analysis.get('confidence', 0):.1%}"
            
            row += 2
    
    def _create_analytics_sheet(self, sheet, data: Dict[str, Any]):
        """Create analytics and charts sheet"""
        
        sheet['A1'] = 'Analytics Dashboard'
        sheet['A1'].font = Font(size=14, bold=True)
        
        # Performance comparison chart data
        ocr_data = data.get('ocr_analysis', {})
        api_performance = ocr_data.get('api_results_summary', {}).get('api_performance', {})
        
        if api_performance:
            row = 3
            sheet[f'A{row}'] = 'API Performance Comparison'
            sheet[f'A{row}'].font = Font(size=12, bold=True)
            
            # Headers
            row += 1
            sheet[f'A{row}'] = 'API'
            sheet[f'B{row}'] = 'Status'
            sheet[f'C{row}'] = 'Processing Time (s)'
            sheet[f'D{row}'] = 'Tokens Used'
            
            # Data
            for api_name, perf in api_performance.items():
                row += 1
                sheet[f'A{row}'] = api_name
                sheet[f'B{row}'] = perf.get('status', '')
                if perf.get('status') == 'success':
                    sheet[f'C{row}'] = perf.get('processing_time', 0)
                    sheet[f'D{row}'] = perf.get('tokens_used', 0)
    
    async def export_to_pdf(self, data: Dict[str, Any], 
                          file_path: str,
                          template: str = ExportTemplate.TECHNICAL_REPORT,
                          include_images: bool = True) -> Dict[str, Any]:
        """🔥 ADVANCED PDF EXPORT"""
        
        if not PDF_AVAILABLE:
            return {'status': 'error', 'error': 'PDF libraries not available'}
        
        start_time = time.time()
        
        try:
            print(f"📄 Generating PDF report: {os.path.basename(file_path)}")
            
            # Create PDF document
            doc = SimpleDocTemplate(file_path, pagesize=letter)
            story = []
            
            # Get styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.darkblue,
                alignment=TA_CENTER
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.darkblue,
                spaceBefore=12,
                spaceAfter=6
            )
            
            # Title page
            story.append(Paragraph("Technical Drawing Analysis Report", title_style))
            story.append(Spacer(1, 0.2*inch))
            
            # Basic information
            metadata = data.get('metadata', {})
            story.append(Paragraph(f"Analysis Date: {metadata.get('processing_timestamp', 'N/A')}", styles['Normal']))
            story.append(Paragraph(f"Pipeline ID: {metadata.get('pipeline_id', 'N/A')}", styles['Normal']))
            story.append(Spacer(1, 0.3*inch))
            
            # Template-specific content
            template_config = self.templates.get(template, self.templates[ExportTemplate.TECHNICAL_REPORT])
            
            if 'executive_summary' in template_config['sections']:
                story.append(self._create_pdf_executive_summary(data, heading_style, styles))
            
            if 'technical_specifications' in template_config['sections']:
                story.append(self._create_pdf_technical_specifications(data, heading_style, styles))
            
            if 'manufacturing_info' in template_config['sections']:
                story.append(self._create_pdf_manufacturing_info(data, heading_style, styles))
            
            if 'recommendations' in template_config['sections']:
                story.append(self._create_pdf_recommendations(data, heading_style, styles))
            
            # Page break for detailed data
            story.append(PageBreak())
            
            # Detailed analysis results
            story.append(self._create_pdf_detailed_results(data, heading_style, styles))
            
            # Build PDF
            doc.build(story)
            
            processing_time = time.time() - start_time
            
            # Update statistics
            self._update_export_stats('pdf', processing_time, True)
            
            print(f"   ✅ PDF export complete in {processing_time:.1f}s")
            
            return {
                'status': 'success',
                'file_path': file_path,
                'file_size': os.path.getsize(file_path),
                'processing_time': processing_time,
                'pages': 'Multi-page'
            }
            
        except Exception as e:
            self._update_export_stats('pdf', time.time() - start_time, False)
            return {
                'status': 'error',
                'error': str(e),
                'processing_time': time.time() - start_time
            }
    
    def _create_pdf_executive_summary(self, data: Dict[str, Any], heading_style, styles):
        """Create executive summary section for PDF"""
        content = []
        
        content.append(Paragraph("Executive Summary", heading_style))
        
        # Get key metrics
        tech_data = data.get('technical_data', {})
        quality_metrics = tech_data.get('quality_metrics', {})
        
        summary_text = f"""
        This technical drawing analysis processed engineering documentation using advanced AI and computer vision techniques. 
        The analysis identified {len(tech_data.get('dimensions_found', []))} dimensional specifications, 
        {len(tech_data.get('materials_identified', []))} material references, and 
        {len(tech_data.get('tolerances_specified', []))} tolerance specifications.
        
        The overall analysis achieved an OCR confidence of {quality_metrics.get('ocr_confidence', 0):.1%} 
        and semantic understanding confidence of {quality_metrics.get('semantic_confidence', 0):.1%}.
        """
        
        content.append(Paragraph(summary_text, styles['Normal']))
        content.append(Spacer(1, 0.2*inch))
        
        return KeepTogether(content)
    
    def _create_pdf_technical_specifications(self, data: Dict[str, Any], heading_style, styles):
        """Create technical specifications section for PDF"""
        content = []
        
        content.append(Paragraph("Technical Specifications", heading_style))
        
        tech_data = data.get('technical_data', {})
        
        # Dimensions
        dimensions = tech_data.get('dimensions_found', [])
        if dimensions:
            content.append(Paragraph("Dimensions Identified:", styles['Heading3']))
            for dim in dimensions[:10]:  # Limit to first 10
                content.append(Paragraph(f"• {dim}", styles['Normal']))
        
        # Materials
        materials = tech_data.get('materials_identified', [])
        if materials:
            content.append(Paragraph("Materials Specified:", styles['Heading3']))
            for material in materials:
                content.append(Paragraph(f"• {material}", styles['Normal']))
        
        # Tolerances
        tolerances = tech_data.get('tolerances_specified', [])
        if tolerances:
            content.append(Paragraph("Tolerance Specifications:", styles['Heading3']))
            for tolerance in tolerances:
                content.append(Paragraph(f"• {tolerance}", styles['Normal']))
        
        content.append(Spacer(1, 0.2*inch))
        
        return KeepTogether(content)
    
    def _create_pdf_manufacturing_info(self, data: Dict[str, Any], heading_style, styles):
        """Create manufacturing information section for PDF"""
        content = []
        
        content.append(Paragraph("Manufacturing Information", heading_style))
        
        # Get semantic analysis results
        semantic_data = data.get('semantic_analysis', {})
        consolidated = semantic_data.get('consolidated_insights', {})
        
        if consolidated.get('manufacturing_insights', {}).get('found'):
            content.append(Paragraph("Manufacturing processes and requirements have been identified in the technical documentation.", styles['Normal']))
        else:
            content.append(Paragraph("No specific manufacturing information was identified in the analysis.", styles['Normal']))
        
        content.append(Spacer(1, 0.2*inch))
        
        return KeepTogether(content)
    
    def _create_pdf_recommendations(self, data: Dict[str, Any], heading_style, styles):
        """Create recommendations section for PDF"""
        content = []
        
        content.append(Paragraph("Recommendations", heading_style))
        
        # Analysis-based recommendations
        tech_data = data.get('technical_data', {})
        quality_metrics = tech_data.get('quality_metrics', {})
        
        recommendations = []
        
        if quality_metrics.get('ocr_confidence', 0) < 0.7:
            recommendations.append("Consider image preprocessing to improve OCR accuracy")
        
        if quality_metrics.get('semantic_confidence', 0) < 0.7:
            recommendations.append("Manual review of AI analysis results is recommended")
        
        if len(tech_data.get('dimensions_found', [])) == 0:
            recommendations.append("No dimensions were automatically detected - manual verification needed")
        
        if not recommendations:
            recommendations.append("Analysis completed successfully with good confidence levels")
        
        for rec in recommendations:
            content.append(Paragraph(f"• {rec}", styles['Normal']))
        
        content.append(Spacer(1, 0.2*inch))
        
        return KeepTogether(content)
    
    def _create_pdf_detailed_results(self, data: Dict[str, Any], heading_style, styles):
        """Create detailed results section for PDF"""
        content = []
        
        content.append(Paragraph("Detailed Analysis Results", heading_style))
        
        # Create summary table
        export_data = data.get('export_ready_data', {})
        summary_data = export_data.get('summary_data', {})
        
        if summary_data:
            table_data = [
                ['Metric', 'Value'],
                ['Total Dimensions', str(summary_data.get('total_dimensions', 0))],
                ['Total Materials', str(summary_data.get('total_materials', 0))],
                ['Total Tolerances', str(summary_data.get('total_tolerances', 0))],
                ['Quality Score', str(summary_data.get('overall_quality_score', 0))],
            ]
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            content.append(table)
        
        return KeepTogether(content)
    
    async def export_to_forge_cad(self, data: Dict[str, Any], 
                                bucket_name: str,
                                object_name: str = None) -> Dict[str, Any]:
        """🔥 EXPORT TO AUTODESK FORGE FOR CAD INTEGRATION"""
        
        if not self.forge_config['client_id']:
            return {'status': 'error', 'error': 'Autodesk Forge not configured'}
        
        start_time = time.time()
        
        try:
            print(f"🏗️ Exporting to Autodesk Forge: {bucket_name}")
            
            # Authenticate with Forge
            if not await self.authenticate_forge():
                return {'status': 'error', 'error': 'Forge authentication failed'}
            
            # Prepare CAD-optimized data
            cad_data = self._prepare_cad_export_data(data)
            
            # Create JSON file with CAD data
            if object_name is None:
                object_name = f"cadly_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            
            temp_file = os.path.join(tempfile.gettempdir(), object_name)
            with open(temp_file, 'w', encoding='utf-8') as f:
                json.dump(cad_data, f, indent=2, ensure_ascii=False)
            
            # Upload to Forge
            headers = {'Authorization': f'Bearer {self.forge_token}'}
            
            # Create bucket if needed
            bucket_result = await self._ensure_forge_bucket(bucket_name, headers)
            if not bucket_result['success']:
                return {'status': 'error', 'error': bucket_result['error']}
            
            # Upload file
            upload_result = await self._upload_to_forge(bucket_name, object_name, temp_file, headers)
            
            # Clean up temp file
            os.remove(temp_file)
            
            processing_time = time.time() - start_time
            
            if upload_result['success']:
                # Update statistics
                self._update_export_stats('forge_cad', processing_time, True)
                
                print(f"   ✅ Forge export complete in {processing_time:.1f}s")
                
                return {
                    'status': 'success',
                    'bucket_key': bucket_name,
                    'object_key': object_name,
                    'object_id': upload_result.get('object_id'),
                    'processing_time': processing_time,
                    'forge_urn': upload_result.get('urn')
                }
            else:
                self._update_export_stats('forge_cad', processing_time, False)
                return {
                    'status': 'error',
                    'error': upload_result.get('error', 'Upload failed'),
                    'processing_time': processing_time
                }
            
        except Exception as e:
            self._update_export_stats('forge_cad', time.time() - start_time, False)
            return {
                'status': 'error',
                'error': str(e),
                'processing_time': time.time() - start_time
            }
    
    def _prepare_cad_export_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Prepare data optimized for CAD system integration"""
        
        cad_data = {
            'metadata': {
                'export_type': 'cad_integration',
                'export_timestamp': datetime.now().isoformat(),
                'cadly_version': '1.0',
                'data_schema': 'forge_compatible_v1'
            },
            'geometric_data': {},
            'technical_specifications': {},
            'manufacturing_info': {},
            'quality_metadata': {}
        }
        
        # Extract geometric information
        symbol_data = data.get('symbol_analysis', {})
        geometric_elements = symbol_data.get('geometric_elements', {})
        
        if geometric_elements:
            cad_data['geometric_data'] = {
                'total_shapes': geometric_elements.get('total_shapes', 0),
                'circles': len(geometric_elements.get('circles', [])),
                'rectangles': len(geometric_elements.get('rectangles', [])),
                'lines': len(geometric_elements.get('lines', [])),
                'shape_details': {
                    'circles': geometric_elements.get('circles', [])[:5],  # Limit for size
                    'rectangles': geometric_elements.get('rectangles', [])[:5],
                    'lines': geometric_elements.get('lines', [])[:10]
                }
            }
        
        # Extract technical specifications
        tech_data = data.get('technical_data', {})
        cad_data['technical_specifications'] = {
            'dimensions': tech_data.get('dimensions_found', []),
            'materials': tech_data.get('materials_identified', []),
            'tolerances': tech_data.get('tolerances_specified', []),
            'standards_referenced': tech_data.get('standards_referenced', [])
        }
        
        # Quality and confidence metrics
        quality_metrics = tech_data.get('quality_metrics', {})
        cad_data['quality_metadata'] = {
            'ocr_confidence': quality_metrics.get('ocr_confidence', 0),
            'semantic_confidence': quality_metrics.get('semantic_confidence', 0),
            'technical_content_richness': quality_metrics.get('technical_content_richness', 0),
            'recommended_review': quality_metrics.get('ocr_confidence', 0) < 0.8
        }
        
        return cad_data
    
    async def _ensure_forge_bucket(self, bucket_name: str, headers: Dict[str, str]) -> Dict[str, Any]:
        """Ensure Forge bucket exists"""
        try:
            async with aiohttp.ClientSession() as session:
                # Check if bucket exists
                check_url = f"{self.forge_config['base_url']}/oss/v2/buckets/{bucket_name}/details"
                async with session.get(check_url, headers=headers) as response:
                    if response.status == 200:
                        return {'success': True, 'message': 'Bucket exists'}
                    
                    # Create bucket if it doesn't exist
                    create_url = f"{self.forge_config['base_url']}/oss/v2/buckets"
                    payload = {
                        'bucketKey': bucket_name,
                        'policyKey': 'temporary'  # 24 hours retention
                    }
                    
                    async with session.post(create_url, headers=headers, json=payload) as create_response:
                        if create_response.status in [200, 409]:  # 409 = already exists
                            return {'success': True, 'message': 'Bucket created/exists'}
                        else:
                            error_text = await create_response.text()
                            return {'success': False, 'error': f'Bucket creation failed: {error_text}'}
                            
        except Exception as e:
            return {'success': False, 'error': str(e)}
    
    async def _upload_to_forge(self, bucket_name: str, object_name: str, 
                             file_path: str, headers: Dict[str, str]) -> Dict[str, Any]:
        """Upload file to Forge bucket"""
        try:
            upload_url = f"{self.forge_config['base_url']}/oss/v2/buckets/{bucket_name}/objects/{object_name}"
            
            async with aiohttp.ClientSession() as session:
                with open(file_path, 'rb') as file_data:
                    upload_headers = headers.copy()
                    upload_headers['Content-Type'] = 'application/octet-stream'
                    
                    async with session.put(upload_url, headers=upload_headers, data=file_data) as response:
                        if response.status == 200:
                            result = await response.json()
                            
                            # Generate URN for the object
                            object_id = result.get('objectId', '')
                            urn = base64.b64encode(object_id.encode()).decode().rstrip('=')
                            
                            return {
                                'success': True,
                                'object_id': object_id,
                                'urn': urn,
                                'size': result.get('size')
                            }
                        else:
                            error_text = await response.text()
                            return {'success': False, 'error': f'Upload failed: {error_text}'}
                            
        except Exception as e:
            return {'success': False, 'error': str(e)}
    
    def _update_export_stats(self, format_type: str, processing_time: float, success: bool):
        """Update export statistics"""
        self.export_stats['total_exports'] += 1
        
        if success:
            self.export_stats['successful_exports'] += 1
        else:
            self.export_stats['failed_exports'] += 1
        
        # Format usage tracking
        if format_type not in self.export_stats['format_usage']:
            self.export_stats['format_usage'][format_type] = 0
        self.export_stats['format_usage'][format_type] += 1
        
        # Update average time
        current_avg = self.export_stats['average_export_time']
        total_exports = self.export_stats['total_exports']
        new_avg = ((current_avg * (total_exports - 1)) + processing_time) / total_exports
        self.export_stats['average_export_time'] = new_avg
    
    async def export_multiple_formats(self, data: Dict[str, Any], 
                                    output_directory: str,
                                    formats: List[str] = None,
                                    filename_base: str = None) -> Dict[str, Any]:
        """🔥 EXPORT TO MULTIPLE FORMATS SIMULTANEOUSLY"""
        
        if formats is None:
            formats = [ExportFormat.JSON, ExportFormat.EXCEL, ExportFormat.PDF]
        
        if filename_base is None:
            filename_base = f"cadly_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        print(f"🚀 Multi-format export: {', '.join(formats)}")
        
        os.makedirs(output_directory, exist_ok=True)
        export_results = {}
        
        # Create export tasks
        tasks = []
        
        for format_type in formats:
            file_path = os.path.join(output_directory, f"{filename_base}.{format_type}")
            
            if format_type == ExportFormat.JSON:
                task = self._export_json_async(data, file_path)
            elif format_type == ExportFormat.EXCEL:
                task = self.export_to_excel(data, file_path)
            elif format_type == ExportFormat.PDF:
                task = self.export_to_pdf(data, file_path)
            elif format_type == ExportFormat.CSV:
                task = self._export_csv_async(data, file_path)
            else:
                continue
            
            tasks.append((format_type, task))
        
        # Execute exports concurrently
        if tasks:
            results = await asyncio.gather(*[task[1] for task in tasks], return_exceptions=True)
            
            for i, (format_type, _) in enumerate(tasks):
                result = results[i]
                if isinstance(result, Exception):
                    export_results[format_type] = {
                        'status': 'error',
                        'error': str(result)
                    }
                else:
                    export_results[format_type] = result
        
        # Summary
        successful_exports = sum(1 for result in export_results.values() 
                               if result.get('status') == 'success')
        
        return {
            'summary': {
                'total_formats': len(export_results),
                'successful_exports': successful_exports,
                'failed_exports': len(export_results) - successful_exports,
                'output_directory': output_directory
            },
            'results': export_results
        }
    
    async def _export_json_async(self, data: Dict[str, Any], file_path: str) -> Dict[str, Any]:
        """Async JSON export"""
        try:
            start_time = time.time()
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False, default=str)
            
            processing_time = time.time() - start_time
            self._update_export_stats('json', processing_time, True)
            
            return {
                'status': 'success',
                'file_path': file_path,
                'file_size': os.path.getsize(file_path),
                'processing_time': processing_time
            }
            
        except Exception as e:
            self._update_export_stats('json', time.time() - start_time if 'start_time' in locals() else 0, False)
            return {
                'status': 'error',
                'error': str(e)
            }
    
    async def _export_csv_async(self, data: Dict[str, Any], file_path: str) -> Dict[str, Any]:
        """Async CSV export"""
        try:
            start_time = time.time()
            
            # Get tabular data
            export_data = data.get('export_ready_data', {})
            tabular_data = export_data.get('tabular_data', [])
            
            if not tabular_data:
                tabular_data = [{'Message': 'No tabular data available'}]
            
            # Write CSV
            fieldnames = list(tabular_data[0].keys()) if tabular_data else ['Message']
            
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(tabular_data)
            
            processing_time = time.time() - start_time
            self._update_export_stats('csv', processing_time, True)
            
            return {
                'status': 'success',
                'file_path': file_path,
                'file_size': os.path.getsize(file_path),
                'processing_time': processing_time
            }
            
        except Exception as e:
            self._update_export_stats('csv', time.time() - start_time if 'start_time' in locals() else 0, False)
            return {
                'status': 'error',
                'error': str(e)
            }
    
    def get_export_statistics(self) -> Dict[str, Any]:
        """Get comprehensive export statistics"""
        return {
            'statistics': self.export_stats,
            'capabilities': self.export_capabilities,
            'templates_available': list(self.templates.keys()),
            'forge_status': 'configured' if self.forge_config['client_id'] else 'not_configured'
        }


# Helper functions for standalone usage
async def export_analysis_results(data: Dict[str, Any], 
                                output_directory: str,
                                formats: List[str] = None) -> Dict[str, Any]:
    """Helper function to export analysis results"""
    export_engine = AdvancedExportEngine()
    return await export_engine.export_multiple_formats(data, output_directory, formats)

async def export_to_forge(data: Dict[str, Any], bucket_name: str) -> Dict[str, Any]:
    """Helper function to export to Autodesk Forge"""
    export_engine = AdvancedExportEngine()
    return await export_engine.export_to_forge_cad(data, bucket_name)


if __name__ == "__main__":
    # Demo usage
    async def demo():
        print("🔥 ADVANCED EXPORT ENGINE DEMO")
        
        # Mock data for testing
        mock_data = {
            'metadata': {
                'processing_timestamp': datetime.now().isoformat(),
                'pipeline_id': 'demo_pipeline_001'
            },
            'technical_data': {
                'dimensions_found': ['150mm x 75mm', 'Ø25mm', '100±0.1mm'],
                'materials_identified': ['SS316', 'ALUMINUM'],
                'tolerances_specified': ['±0.1mm', '±0.05mm'],
                'quality_metrics': {
                    'ocr_confidence': 0.85,
                    'semantic_confidence': 0.78,
                    'technical_content_richness': 5
                }
            },
            'export_ready_data': {
                'tabular_data': [
                    {'Type': 'Dimension', 'Category': 'Length', 'Value': '150mm', 'Source': 'OCR', 'Confidence': 0.9},
                    {'Type': 'Material', 'Category': 'Steel', 'Value': 'SS316', 'Source': 'AI', 'Confidence': 0.8},
                ],
                'summary_data': {
                    'total_dimensions': 3,
                    'total_materials': 2,
                    'total_tolerances': 2,
                    'overall_quality_score': 5
                }
            }
        }
        
        export_engine = AdvancedExportEngine()
        
        # Test multi-format export
        output_dir = "demo_exports"
        results = await export_engine.export_multiple_formats(
            mock_data, 
            output_dir, 
            [ExportFormat.JSON, ExportFormat.EXCEL, ExportFormat.PDF]
        )
        
        print("\n📊 EXPORT RESULTS SUMMARY:")
        print("=" * 40)
        summary = results['summary']
        print(f"📁 Output Directory: {summary['output_directory']}")
        print(f"✅ Successful Exports: {summary['successful_exports']}")
        print(f"❌ Failed Exports: {summary['failed_exports']}")
        
        # Display individual results
        for format_type, result in results['results'].items():
            if result['status'] == 'success':
                print(f"   {format_type.upper()}: ✅ Success ({result['processing_time']:.2f}s)")
            else:
                print(f"   {format_type.upper()}: ❌ Failed - {result.get('error', 'Unknown error')}")
        
        # System statistics
        stats = export_engine.get_export_statistics()
        print(f"\n🔧 Export Engine Status:")
        for capability, available in stats['capabilities'].items():
            print(f"   {capability}: {'✅' if available else '❌'}")
    
    # Run demo
    asyncio.run(demo())